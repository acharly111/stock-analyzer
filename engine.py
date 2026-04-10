import yfinance as yf
import pandas as pd
from openpyxl import load_workbook
import json
import os


EARNINGS_CACHE_FILE = "earnings_cache.json"


# ---------- Earnings cache ----------
def load_earnings_cache():
    if not os.path.exists(EARNINGS_CACHE_FILE):
        return {}
    try:
        with open(EARNINGS_CACHE_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
            return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def save_earnings_cache(cache):
    try:
        with open(EARNINGS_CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump(cache, f, indent=2)
    except Exception:
        pass


# ---------- Helpers ----------
def calculate_atr(data, period=14):
    high = data["High"]
    low = data["Low"]
    close = data["Close"]

    tr1 = high - low
    tr2 = (high - close.shift()).abs()
    tr3 = (low - close.shift()).abs()

    true_range = pd.concat([tr1, tr2, tr3], axis=1).max(axis=1)
    atr = true_range.ewm(alpha=1 / period, min_periods=period).mean()
    return atr


def classify_risk(atr_pct):
    if atr_pct >= 0.06:
        return "Very High"
    elif atr_pct >= 0.045:
        return "High"
    elif atr_pct >= 0.03:
        return "Mid"
    elif atr_pct >= 0.015:
        return "Low"
    else:
        return "Very Low"


def to_scalar(x):
    if isinstance(x, pd.Series):
        return x.iloc[0]
    return x


def get_close_on_date(data, target_date):
    if target_date is None or data.empty:
        return None

    target_date = pd.Timestamp(target_date).normalize()
    matches = data.loc[data.index.normalize() == target_date]

    if matches.empty:
        return None

    value = matches["Close"].iloc[-1]
    value = to_scalar(value)

    if pd.isna(value):
        return None

    return round(float(value), 2)


def calc_pct_change(start_value, end_value):
    if start_value is None or end_value is None:
        return None
    if start_value == 0:
        return None
    return round(((end_value / start_value) - 1) * 100, 2)


def _normalize_earnings_date(value):
    if value is None:
        return None
    try:
        ts = pd.Timestamp(value)
        if pd.isna(ts):
            return None
        if ts.tzinfo is not None:
            ts = ts.tz_convert(None)
        return ts.normalize()
    except Exception:
        return None


# ---------- Earnings lookup ----------
def _fetch_next_earnings_date_from_yahoo(symbol):
    try:
        ticker = yf.Ticker(symbol)

        # First try get_earnings_dates
        try:
            edf = ticker.get_earnings_dates(limit=12)
            if edf is not None and not edf.empty:
                today = pd.Timestamp.today().normalize()
                dates = []

                for idx in edf.index:
                    ts = _normalize_earnings_date(idx)
                    if ts is not None and ts >= today:
                        dates.append(ts)

                if dates:
                    return min(dates).strftime("%Y-%m-%d")
        except Exception:
            pass

        # Fallback to calendar
        try:
            cal = ticker.calendar

            if isinstance(cal, dict):
                earnings_value = cal.get("Earnings Date")

                if isinstance(earnings_value, (list, tuple)):
                    today = pd.Timestamp.today().normalize()
                    dates = []
                    for item in earnings_value:
                        ts = _normalize_earnings_date(item)
                        if ts is not None and ts >= today:
                            dates.append(ts)
                    if dates:
                        return min(dates).strftime("%Y-%m-%d")
                else:
                    ts = _normalize_earnings_date(earnings_value)
                    if ts is not None and ts >= today:
                        return ts.strftime("%Y-%m-%d")

            if isinstance(cal, pd.DataFrame) and not cal.empty:
                today = pd.Timestamp.today().normalize()
                dates = []

                for idx in cal.index:
                    ts = _normalize_earnings_date(idx)
                    if ts is not None and ts >= today:
                        dates.append(ts)

                if not dates:
                    for col in cal.columns:
                        for item in cal[col].tolist():
                            ts = _normalize_earnings_date(item)
                            if ts is not None and ts >= today:
                                dates.append(ts)

                if dates:
                    return min(dates).strftime("%Y-%m-%d")
        except Exception:
            pass

    except Exception:
        pass

    return None


def get_next_earnings_date(symbol):
    """
    Reuse cached date unless it is within 21 days or already passed.
    If Yahoo refresh fails, reuse old cached value.
    If nothing is available, return 'N/A'.
    """
    cache = load_earnings_cache()
    today = pd.Timestamp.today().normalize()

    cached_date_str = cache.get(symbol)

    if cached_date_str:
        try:
            cached_ts = pd.Timestamp(cached_date_str).normalize()
            days_to_cached = (cached_ts - today).days

            if days_to_cached > 21:
                return cached_date_str
        except Exception:
            pass

    new_date = _fetch_next_earnings_date_from_yahoo(symbol)

    if new_date:
        cache[symbol] = new_date
        save_earnings_cache(cache)
        return new_date

    if cached_date_str:
        return cached_date_str

    return "N/A"


# ---------- Core analysis ----------
def analyze_stocks(tickers, start_date=None, low_date=None):
    results = []

    for ticker in tickers:
        try:
            data = yf.download(
                ticker,
                period="2y",
                interval="1d",
                progress=False,
                auto_adjust=False,
                group_by="column"
            )

            next_earnings_date = get_next_earnings_date(ticker)

            if data.empty:
                results.append({
                    "symbol": ticker,
                    "price": None,
                    "ATR_14": None,
                    "ATR_pct": None,
                    "risk_level": "No Data",
                    "Close_Start_Date": None,
                    "Close_Low_Date": None,
                    "Current_Price": None,
                    "Change_%_Start_to_Today": None,
                    "Change_%_Low_to_Today": None,
                    "Strong_vs_SPY": None,
                    "SMA_20": None,
                    "SMA_50": None,
                    "SMA_100": None,
                    "SMA_150": None,
                    "SMA_200": None,
                    "RSI_14": None,
                    "RSI_MA_14": None,
                    "RSI_dist_%_from_RSI_MA_14": None,
                    "Price_dist_%_from_SMA_20": None,
                    "Price_dist_%_from_SMA_50": None,
                    "Price_dist_%_from_SMA_100": None,
                    "Price_dist_%_from_SMA_150": None,
                    "Price_dist_%_from_SMA_200": None,
                    "next_earnings_date": next_earnings_date,
                    "BUY/SELL signal": None,
                })
                continue

            if isinstance(data.columns, pd.MultiIndex):
                data.columns = data.columns.get_level_values(0)

            data["ATR_14"] = calculate_atr(data)

            for sma in [20, 50, 100, 150, 200]:
                data[f"SMA_{sma}"] = data["Close"].rolling(sma).mean()

            delta = data["Close"].diff()
            gain = delta.clip(lower=0)
            loss = -delta.clip(upper=0)

            avg_gain = gain.ewm(alpha=1 / 14, min_periods=14).mean()
            avg_loss = loss.ewm(alpha=1 / 14, min_periods=14).mean()

            rs = avg_gain / avg_loss
            data["RSI_14"] = 100 - (100 / (1 + rs))
            data["RSI_MA_14"] = data["RSI_14"].rolling(14).mean()

            price = to_scalar(data["Close"].iloc[-1])
            atr = to_scalar(data["ATR_14"].iloc[-1])

            sma_vals = {}
            for sma in [20, 50, 100, 150, 200]:
                sma_val = to_scalar(data[f"SMA_{sma}"].iloc[-1])
                sma_vals[sma] = round(float(sma_val), 2) if pd.notna(sma_val) else None

            rsi_14 = to_scalar(data["RSI_14"].iloc[-1])
            rsi_ma_14 = to_scalar(data["RSI_MA_14"].iloc[-1])

            close_start = get_close_on_date(data, start_date)
            close_low = get_close_on_date(data, low_date)
            current_price = round(float(price), 2) if pd.notna(price) else None

            if pd.isna(price) or pd.isna(atr) or float(price) == 0:
                results.append({
                    "symbol": ticker,
                    "price": None,
                    "ATR_14": None,
                    "ATR_pct": None,
                    "risk_level": "No Data",
                    "Close_Start_Date": close_start,
                    "Close_Low_Date": close_low,
                    "Current_Price": current_price,
                    "Change_%_Start_to_Today": None,
                    "Change_%_Low_to_Today": None,
                    "Strong_vs_SPY": None,
                    "SMA_20": sma_vals[20],
                    "SMA_50": sma_vals[50],
                    "SMA_100": sma_vals[100],
                    "SMA_150": sma_vals[150],
                    "SMA_200": sma_vals[200],
                    "RSI_14": round(float(rsi_14), 2) if pd.notna(rsi_14) else None,
                    "RSI_MA_14": round(float(rsi_ma_14), 2) if pd.notna(rsi_ma_14) else None,
                    "RSI_dist_%_from_RSI_MA_14": None,
                    "Price_dist_%_from_SMA_20": None,
                    "Price_dist_%_from_SMA_50": None,
                    "Price_dist_%_from_SMA_100": None,
                    "Price_dist_%_from_SMA_150": None,
                    "Price_dist_%_from_SMA_200": None,
                    "next_earnings_date": next_earnings_date,
                    "BUY/SELL signal": None,
                })
                continue

            price = float(price)
            atr = float(atr)
            atr_pct = atr / price
            risk = classify_risk(atr_pct)

            price_dist = {}
            for sma in [20, 50, 100, 150, 200]:
                if sma_vals[sma] not in [None, 0]:
                    price_dist[sma] = round(((price / sma_vals[sma]) - 1) * 100, 2)
                else:
                    price_dist[sma] = None

            rsi_dist = None
            if pd.notna(rsi_14) and pd.notna(rsi_ma_14) and float(rsi_ma_14) != 0:
                rsi_dist = round(((float(rsi_14) / float(rsi_ma_14)) - 1) * 100, 2)

            results.append({
                "symbol": ticker,
                "price": round(price, 2),
                "ATR_14": round(atr, 2),
                "ATR_pct": round(atr_pct * 100, 2),
                "risk_level": risk,
                "Close_Start_Date": close_start,
                "Close_Low_Date": close_low,
                "Current_Price": current_price,
                "Change_%_Start_to_Today": None,
                "Change_%_Low_to_Today": None,
                "Strong_vs_SPY": None,
                "SMA_20": sma_vals[20],
                "SMA_50": sma_vals[50],
                "SMA_100": sma_vals[100],
                "SMA_150": sma_vals[150],
                "SMA_200": sma_vals[200],
                "RSI_14": round(float(rsi_14), 2) if pd.notna(rsi_14) else None,
                "RSI_MA_14": round(float(rsi_ma_14), 2) if pd.notna(rsi_ma_14) else None,
                "RSI_dist_%_from_RSI_MA_14": rsi_dist,
                "Price_dist_%_from_SMA_20": price_dist[20],
                "Price_dist_%_from_SMA_50": price_dist[50],
                "Price_dist_%_from_SMA_100": price_dist[100],
                "Price_dist_%_from_SMA_150": price_dist[150],
                "Price_dist_%_from_SMA_200": price_dist[200],
                "next_earnings_date": next_earnings_date,
                "BUY/SELL signal": None,
            })

        except Exception:
            results.append({
                "symbol": ticker,
                "price": None,
                "ATR_14": None,
                "ATR_pct": None,
                "risk_level": "Error",
                "Close_Start_Date": None,
                "Close_Low_Date": None,
                "Current_Price": None,
                "Change_%_Start_to_Today": None,
                "Change_%_Low_to_Today": None,
                "Strong_vs_SPY": None,
                "SMA_20": None,
                "SMA_50": None,
                "SMA_100": None,
                "SMA_150": None,
                "SMA_200": None,
                "RSI_14": None,
                "RSI_MA_14": None,
                "RSI_dist_%_from_RSI_MA_14": None,
                "Price_dist_%_from_SMA_20": None,
                "Price_dist_%_from_SMA_50": None,
                "Price_dist_%_from_SMA_100": None,
                "Price_dist_%_from_SMA_150": None,
                "Price_dist_%_from_SMA_200": None,
                "next_earnings_date": "N/A",
                "BUY/SELL signal": None,
            })

    return pd.DataFrame(results)


def get_spy_row(start_date=None, low_date=None):
    try:
        data = yf.download(
            "SPY",
            period="2y",
            interval="1d",
            progress=False,
            auto_adjust=False,
            group_by="column"
        )

        if data.empty:
            raise ValueError("SPY data empty")

        if isinstance(data.columns, pd.MultiIndex):
            data.columns = data.columns.get_level_values(0)

        for sma in [20, 50, 100, 150, 200]:
            data[f"SMA_{sma}"] = data["Close"].rolling(sma).mean()

        delta = data["Close"].diff()
        gain = delta.clip(lower=0)
        loss = -delta.clip(upper=0)
        avg_gain = gain.ewm(alpha=1 / 14, min_periods=14).mean()
        avg_loss = loss.ewm(alpha=1 / 14, min_periods=14).mean()
        rs = avg_gain / avg_loss
        data["RSI_14"] = 100 - (100 / (1 + rs))
        data["RSI_MA_14"] = data["RSI_14"].rolling(14).mean()

        price = float(to_scalar(data["Close"].dropna().iloc[-1]))

        sma_vals = {}
        for sma in [20, 50, 100, 150, 200]:
            series = data[f"SMA_{sma}"].dropna()
            sma_vals[sma] = round(float(to_scalar(series.iloc[-1])), 2) if not series.empty else None

        rsi_14_series = data["RSI_14"].dropna()
        rsi_ma_14_series = data["RSI_MA_14"].dropna()

        rsi_14 = round(float(to_scalar(rsi_14_series.iloc[-1])), 2) if not rsi_14_series.empty else None
        rsi_ma_14 = round(float(to_scalar(rsi_ma_14_series.iloc[-1])), 2) if not rsi_ma_14_series.empty else None

        rsi_dist = None
        if rsi_14 is not None and rsi_ma_14 not in [None, 0]:
            rsi_dist = round(((rsi_14 / rsi_ma_14) - 1) * 100, 2)

        price_dist = {}
        for sma in [20, 50, 100, 150, 200]:
            if sma_vals[sma] not in [None, 0]:
                price_dist[sma] = round(((price / sma_vals[sma]) - 1) * 100, 2)
            else:
                price_dist[sma] = None

        close_start = get_close_on_date(data, start_date)
        close_low = get_close_on_date(data, low_date)

        return {
            "symbol": "SPY",
            "price": round(price, 2),
            "ATR_14": None,
            "ATR_pct": None,
            "risk_level": "Benchmark",
            "Close_Start_Date": close_start,
            "Close_Low_Date": close_low,
            "Current_Price": round(price, 2),
            "Change_%_Start_to_Today": None,
            "Change_%_Low_to_Today": None,
            "Strong_vs_SPY": None,
            "SMA_20": sma_vals[20],
            "SMA_50": sma_vals[50],
            "SMA_100": sma_vals[100],
            "SMA_150": sma_vals[150],
            "SMA_200": sma_vals[200],
            "RSI_14": rsi_14,
            "RSI_MA_14": rsi_ma_14,
            "RSI_dist_%_from_RSI_MA_14": rsi_dist,
            "Price_dist_%_from_SMA_20": price_dist[20],
            "Price_dist_%_from_SMA_50": price_dist[50],
            "Price_dist_%_from_SMA_100": price_dist[100],
            "Price_dist_%_from_SMA_150": price_dist[150],
            "Price_dist_%_from_SMA_200": price_dist[200],
            "next_earnings_date": "N/A",
            "BUY/SELL signal": None,
        }

    except Exception:
        return {
            "symbol": "SPY",
            "price": None,
            "ATR_14": None,
            "ATR_pct": None,
            "risk_level": "Benchmark",
            "Close_Start_Date": None,
            "Close_Low_Date": None,
            "Current_Price": None,
            "Change_%_Start_to_Today": None,
            "Change_%_Low_to_Today": None,
            "Strong_vs_SPY": None,
            "SMA_20": None,
            "SMA_50": None,
            "SMA_100": None,
            "SMA_150": None,
            "SMA_200": None,
            "RSI_14": None,
            "RSI_MA_14": None,
            "RSI_dist_%_from_RSI_MA_14": None,
            "Price_dist_%_from_SMA_20": None,
            "Price_dist_%_from_SMA_50": None,
            "Price_dist_%_from_SMA_100": None,
            "Price_dist_%_from_SMA_150": None,
            "Price_dist_%_from_SMA_200": None,
            "next_earnings_date": "N/A",
            "BUY/SELL signal": None,
        }


def fill_gui_columns(
    df,
    selected_sma=200,
    buy_rsi_dist_max=-10,
    buy_dist_selected_sma_max=10,
    sell_rsi_min=70,
    sell_dist_selected_sma_min=40
):
    df = df.copy()

    df["Change_%_Start_to_Today"] = df.apply(
        lambda row: calc_pct_change(row["Close_Start_Date"], row["Current_Price"]),
        axis=1
    )

    df["Change_%_Low_to_Today"] = df.apply(
        lambda row: calc_pct_change(row["Close_Low_Date"], row["Current_Price"]),
        axis=1
    )

    selected_dist_col = f"Price_dist_%_from_SMA_{selected_sma}"

    spy_rows = df[df["symbol"] == "SPY"]
    if not spy_rows.empty:
        spy_change_start = spy_rows.iloc[0]["Change_%_Start_to_Today"]
        spy_change_low = spy_rows.iloc[0]["Change_%_Low_to_Today"]

        def strong_vs_spy(row):
            if row["symbol"] == "SPY":
                return "Benchmark"
            if pd.isna(row["Change_%_Start_to_Today"]) or pd.isna(row["Change_%_Low_to_Today"]):
                return ""
            if pd.isna(spy_change_start) or pd.isna(spy_change_low):
                return ""
            if row["Change_%_Start_to_Today"] > spy_change_start and row["Change_%_Low_to_Today"] > spy_change_low:
                return "Strong"
            return ""

        df["Strong_vs_SPY"] = df.apply(strong_vs_spy, axis=1)

    def signal(row):
        if row["symbol"] == "SPY":
            return "Benchmark"

        sma_value = row.get(f"SMA_{selected_sma}")
        sma_dist = row.get(selected_dist_col)

        buy_ok = (
            pd.notna(row.get("RSI_dist_%_from_RSI_MA_14")) and
            pd.notna(row.get("price")) and
            pd.notna(sma_value) and
            pd.notna(sma_dist) and
            row["RSI_dist_%_from_RSI_MA_14"] < buy_rsi_dist_max and
            row["price"] > sma_value and
            sma_dist < buy_dist_selected_sma_max
        )

        sell_ok = (
            pd.notna(row.get("RSI_14")) and
            pd.notna(sma_dist) and
            row["RSI_14"] >= sell_rsi_min and
            sma_dist > sell_dist_selected_sma_min
        )

        if buy_ok:
            return "BUY"
        if sell_ok:
            return "SELL"
        return ""

    df["BUY/SELL signal"] = df.apply(signal, axis=1)
    return df


def build_output(
    tickers,
    selected_sma=200,
    buy_rsi_dist_max=-10,
    buy_dist_selected_sma_max=10,
    sell_rsi_min=70,
    sell_dist_selected_sma_min=40,
    start_date=None,
    low_date=None
):
    df = analyze_stocks(tickers, start_date=start_date, low_date=low_date)
    spy_row = get_spy_row(start_date=start_date, low_date=low_date)
    df = pd.concat([df, pd.DataFrame([spy_row])], ignore_index=True)

    df = fill_gui_columns(
        df,
        selected_sma=selected_sma,
        buy_rsi_dist_max=buy_rsi_dist_max,
        buy_dist_selected_sma_max=buy_dist_selected_sma_max,
        sell_rsi_min=sell_rsi_min,
        sell_dist_selected_sma_min=sell_dist_selected_sma_min
    )

    return df


def apply_excel_formulas(
    excel_file,
    selected_sma=200,
    buy_rsi_dist_max=-10,
    buy_dist_selected_sma_max=10,
    sell_rsi_min=70,
    sell_dist_selected_sma_min=40
):
    wb = load_workbook(excel_file)
    ws = wb.active

    headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    last_row = ws.max_row
    spy_row_num = last_row

    ws["U1"] = "Parameter"
    ws["V1"] = "Value"
    ws["U2"] = "Selected SMA"
    ws["V2"] = selected_sma
    ws["U3"] = "BUY: RSI distance max (%)"
    ws["V3"] = buy_rsi_dist_max
    ws["U4"] = f"BUY: Price dist from SMA{selected_sma} max (%)"
    ws["V4"] = buy_dist_selected_sma_max
    ws["U5"] = "SELL: RSI min"
    ws["V5"] = sell_rsi_min
    ws["U6"] = f"SELL: Price dist from SMA{selected_sma} min (%)"
    ws["V6"] = sell_dist_selected_sma_min

    if "Change_%_Start_to_Today" in headers:
        c = ws.cell(1, headers["Change_%_Start_to_Today"]).column_letter
        s_col = ws.cell(1, headers["Close_Start_Date"]).column_letter
        t_col = ws.cell(1, headers["Current_Price"]).column_letter
        for r in range(2, last_row + 1):
            ws[f"{c}{r}"] = f'=IF(OR({s_col}{r}="",{t_col}{r}="",{s_col}{r}=0),"",(({t_col}{r}/{s_col}{r})-1)*100)'

    if "Change_%_Low_to_Today" in headers:
        c = ws.cell(1, headers["Change_%_Low_to_Today"]).column_letter
        l_col = ws.cell(1, headers["Close_Low_Date"]).column_letter
        t_col = ws.cell(1, headers["Current_Price"]).column_letter
        for r in range(2, last_row + 1):
            ws[f"{c}{r}"] = f'=IF(OR({l_col}{r}="",{t_col}{r}="",{l_col}{r}=0),"",(({t_col}{r}/{l_col}{r})-1)*100)'

    if "Strong_vs_SPY" in headers and "Change_%_Start_to_Today" in headers and "Change_%_Low_to_Today" in headers:
        kcol = ws.cell(1, headers["Strong_vs_SPY"]).column_letter
        scol = ws.cell(1, headers["Change_%_Start_to_Today"]).column_letter
        lcol = ws.cell(1, headers["Change_%_Low_to_Today"]).column_letter
        for r in range(2, last_row + 1):
            if r == spy_row_num:
                ws[f"{kcol}{r}"] = "Benchmark"
            else:
                ws[f"{kcol}{r}"] = (
                    f'=IF(OR({scol}{r}="",{lcol}{r}="",{scol}{spy_row_num}="",{lcol}{spy_row_num}=""),"",'
                    f'IF(AND({scol}{r}>{scol}{spy_row_num},{lcol}{r}>{lcol}{spy_row_num}),"Strong",""))'
                )

    if "BUY/SELL signal" in headers:
        scol = ws.cell(1, headers["BUY/SELL signal"]).column_letter
        pcol = ws.cell(1, headers["RSI_dist_%_from_RSI_MA_14"]).column_letter
        ncol = ws.cell(1, headers["RSI_14"]).column_letter
        price_col = ws.cell(1, headers["price"]).column_letter
        sma_col = ws.cell(1, headers[f"SMA_{selected_sma}"]).column_letter
        dist_col = ws.cell(1, headers[f"Price_dist_%_from_SMA_{selected_sma}"]).column_letter

        for r in range(2, last_row + 1):
            if r == spy_row_num:
                ws[f"{scol}{r}"] = "Benchmark"
            else:
                ws[f"{scol}{r}"] = (
                    f'=IF(OR({pcol}{r}="",{ncol}{r}="",{price_col}{r}="",{sma_col}{r}="",{dist_col}{r}=""),"",'
                    f'IF(AND({pcol}{r}<$V$3,{price_col}{r}>{sma_col}{r},{dist_col}{r}<$V$4),"BUY",'
                    f'IF(AND({ncol}{r}>=$V$5,{dist_col}{r}>$V$6),"SELL","")))'
                )

    wb.save(excel_file)


def save_outputs(
    df,
    excel_file="risk_analysis.xlsx",
    txt_file="risk_analysis.txt",
    selected_sma=200,
    buy_rsi_dist_max=-10,
    buy_dist_selected_sma_max=10,
    sell_rsi_min=70,
    sell_dist_selected_sma_min=40
):
    df.to_excel(excel_file, index=False)
    apply_excel_formulas(
        excel_file,
        selected_sma=selected_sma,
        buy_rsi_dist_max=buy_rsi_dist_max,
        buy_dist_selected_sma_max=buy_dist_selected_sma_max,
        sell_rsi_min=sell_rsi_min,
        sell_dist_selected_sma_min=sell_dist_selected_sma_min
    )

    with open(txt_file, "w", encoding="utf-8") as f:
        f.write(df.to_string(index=False))